import torch
import time
import numpy as np
import onnxruntime

from openpyxl.utils import column_index_from_string
from openpyxl.styles import Alignment


class EarlyStopper:
    def __init__(self, patience=1, min_delta=0):
        self.patience = patience
        self.min_delta = min_delta
        self.counter = 0
        self.min_validation_loss = float("inf")

    def early_stop(self, validation_loss):
        if validation_loss < self.min_validation_loss:
            self.min_validation_loss = validation_loss
            self.counter = 0
        elif validation_loss > (self.min_validation_loss + self.min_delta):
            self.counter += 1
            if self.counter >= self.patience:
                return True
        return False


# @profile
def benchmark(model_path, batch_size=1, use_gpu=True, is_profile=False):
    # Create an ONNX Runtime session with GPU as the execution provider
    options = onnxruntime.SessionOptions()
    options.graph_optimization_level = onnxruntime.GraphOptimizationLevel.ORT_ENABLE_ALL
    options.enable_profiling = True if is_profile else False

    providers = ["CPUExecutionProvider"]
    if use_gpu and torch.cuda.is_available():
        providers.insert(0, "CUDAExecutionProvider")

    sess = onnxruntime.InferenceSession(
        model_path, providers=providers, sess_options=options
    )

    runs = 10
    device_name = "cpu" if not use_gpu else "cuda"
    device_index = 0
    run_time = []

    # Create an I/O binding
    io_binding = sess.io_binding()

    for i in range(runs):
        for in_tensor in sess.get_inputs():
            # Check the input and output names and shapes of the model
            input_name = in_tensor.name
            input_shape = [batch_size] + in_tensor.shape[1:]
            # print(f"Input name: {input_name}, Input shape: {input_shape}")

            # input_data = torch.empty(
            #     tuple(input_shape), device=device_name
            # ).contiguous()

            input_data = np.random.rand(*input_shape).astype(np.float32)

            X_ortvalue = onnxruntime.OrtValue.ortvalue_from_numpy(
                input_data, device_name, device_index
            )

            # Bind input data to the model
            io_binding.bind_input(
                name=input_name,
                device_type=device_name,
                device_id=device_index,
                element_type=input_data.dtype,
                shape=X_ortvalue.shape(),
                buffer_ptr=X_ortvalue.data_ptr(),
            )

        for out_tensor in sess.get_outputs():
            # Check the input and output names and shapes of the model
            output_name = out_tensor.name
            output_shape = [batch_size] + out_tensor.shape[1:]
            # print(f"Output name: {output_name}, Output shape: {output_shape}")

            # Bind output buffer
            output_data = np.random.rand(*output_shape).astype(np.float32)

            Y_ortvalue = onnxruntime.OrtValue.ortvalue_from_numpy(
                output_data, device_name, device_index
            )

            io_binding.bind_output(
                name=output_name,
                device_type=device_name,
                device_id=device_index,
                element_type=output_data.dtype,
                shape=Y_ortvalue.shape(),
                buffer_ptr=Y_ortvalue.data_ptr(),
            )

        # Run inference
        start = time.time()

        sess.run_with_iobinding(io_binding)

        end = time.time() - start
        run_time.append(end)

    latency = round(np.mean(run_time) * 1000, 2)

    prof_file = sess.end_profiling()

    del sess
    del io_binding

    return latency, prof_file


def extract_analytic_to_excel(workbook, data, title):
    worksheet = workbook.create_sheet(title=f"{title}%_quantized_layers")
    alignment = Alignment(horizontal="center", vertical="center")

    # Add headers to the sheet
    worksheet["A1"] = "Configuration Index"
    # worksheet.merge_cells("A1:A2")

    # worksheet["B1"] = ""
    # worksheet["C1"] = ""
    # worksheet.merge_cells("B1:C1")
    # worksheet["B1"] = "Latency (s)"
    worksheet["B1"] = "FP32 Latency (s)"
    worksheet["C1"] = "INT8 Latency (s)"

    worksheet["D1"] = "Latency Reduction (s)"
    # worksheet.merge_cells("D1:D2")

    worksheet["E1"] = "Ratio Latency Reduction (%)"
    # worksheet.merge_cells("E1:E2")

    # worksheet["F1"] = ""
    # worksheet["G1"] = ""
    # worksheet.merge_cells("F1:G1")
    # worksheet["F1"] = "Accuracy (%)"
    worksheet["F1"] = "FP32 Accuracy (%)"
    worksheet["G1"] = "INT8 Accuracy (%)"

    worksheet["H1"] = "Accuracy Loss (%)"
    # worksheet.merge_cells("H1:H2")

    worksheet["I1"] = "Ratio Accuracy Loss (%)"
    # worksheet.merge_cells("I1:I2")

    worksheet["J1"] = "Number of QuantizeLinear"
    # worksheet.merge_cells("J1:J2")

    worksheet["K1"] = "Number of DequantizeLinear"
    # worksheet.merge_cells("K1:K2")

    worksheet["L1"] = "Parameter of Weight Non-Quantize"
    # worksheet.merge_cells("L1:L2")

    worksheet["M1"] = "Parameter of Weight Quantize"
    # worksheet.merge_cells("M1:M2")

    worksheet["N1"] = "Number of Act Non-Quantize"
    # worksheet.merge_cells("N1:N2")

    worksheet["O1"] = "Number of Act Quantize"
    # worksheet.merge_cells("O1:O2")

    worksheet["P1"] = "Parameter First Weight Quantize"
    # worksheet.merge_cells("P1:P2")

    worksheet["Q1"] = "Parameter Last Weight Quantize"
    # worksheet.merge_cells("Q1:Q2")

    # worksheet["R1"] = ""
    # worksheet["S1"] = ""
    # worksheet.merge_cells("R1:S1")
    # worksheet["R1"] = "Power (W)"
    worksheet["R1"] = "FP32 Power (Watt)"
    worksheet["S1"] = "INT8 Power (Watt)"

    # worksheet["T1"] = ""
    # worksheet["U1"] = ""
    # worksheet.merge_cells("T1:U1")
    # worksheet["T1"] = "Energy (W)"
    worksheet["T1"] = "FP32 Energy (Watt x ms)"
    worksheet["U1"] = "INT8 Energy (Watt x ms)"

    row_index = 3

    for i, d in enumerate(data):
        worksheet.cell(
            row=row_index, column=column_index_from_string("A"), value=str(i)
        )
        worksheet.cell(
            row=row_index,
            column=column_index_from_string("B"),
            value=d["FP32_latency"],
        )
        worksheet.cell(
            row=row_index,
            column=column_index_from_string("C"),
            value=d["INT8_latency"],
        )
        worksheet.cell(
            row=row_index,
            column=column_index_from_string("D"),
            value=d["latency_redution"],
        )
        worksheet.cell(
            row=row_index,
            column=column_index_from_string("E"),
            value=d["ratio_latency_redution"],
        )
        worksheet.cell(
            row=row_index,
            column=column_index_from_string("F"),
            value=d["FP32_accuracy"],
        )
        worksheet.cell(
            row=row_index,
            column=column_index_from_string("G"),
            value=d["INT8_accuracy"],
        )
        worksheet.cell(
            row=row_index,
            column=column_index_from_string("H"),
            value=d["accuracy_loss"],
        )
        worksheet.cell(
            row=row_index,
            column=column_index_from_string("I"),
            value=d["ratio_accuracy_loss"],
        )
        worksheet.cell(
            row=row_index,
            column=column_index_from_string("J"),
            value=d["nb_qlinear"],
        )
        worksheet.cell(
            row=row_index,
            column=column_index_from_string("K"),
            value=d["nb_dqlinear"],
        )
        worksheet.cell(
            row=row_index,
            column=column_index_from_string("L"),
            value=d["param_weight_non_quantized"],
        )
        worksheet.cell(
            row=row_index,
            column=column_index_from_string("M"),
            value=d["param_weight_quantized"],
        )
        worksheet.cell(
            row=row_index,
            column=column_index_from_string("N"),
            value=d["nb_act_non_quantized"],
        )
        worksheet.cell(
            row=row_index,
            column=column_index_from_string("O"),
            value=d["nb_act_quantized"],
        )
        worksheet.cell(
            row=row_index,
            column=column_index_from_string("P"),
            value=d["param_first_quantized"],
        )
        worksheet.cell(
            row=row_index,
            column=column_index_from_string("Q"),
            value=d["param_last_quantized"],
        )
        worksheet.cell(
            row=row_index,
            column=column_index_from_string("R"),
            value=d["FP32_power"],
        )
        worksheet.cell(
            row=row_index,
            column=column_index_from_string("S"),
            value=d["INT8_power"],
        )
        worksheet.cell(
            row=row_index,
            column=column_index_from_string("T"),
            value=d["FP32_energy"],
        )
        worksheet.cell(
            row=row_index,
            column=column_index_from_string("U"),
            value=d["INT8_energy"],
        )

        row_index += 1

    for row in worksheet.iter_rows(
        min_row=1, max_row=row_index, min_col=1, max_col=100
    ):
        for cell in row:
            cell.alignment = alignment
